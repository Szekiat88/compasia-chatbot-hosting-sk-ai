from _ai_config import PROVIDER_PRIMARY
"""
KB Test Runner — simulates WhatsApp messages.

For every question in data/Test_Cases_MainDB_new.xlsx:
  - Sends it through search() exactly as a real WhatsApp message would.
  - Captures the reply that would be sent back to the customer.
  - Saves: question, matched keyword, WhatsApp reply, pass/fail, latency.

Every row is written to Excel immediately after completing — a crash
mid-run never loses finished work.

If the Results sheet already has rows from a previous run, those test
cases are SKIPPED automatically — only the remaining ones are run.
Use --restart to clear all previous results and run from scratch.

Usage:
    python run_kb_tests.py                   # resume / run all remaining
    python run_kb_tests.py --restart         # clear results and run all from scratch
    python run_kb_tests.py --limit 20        # quick smoke-test
    python run_kb_tests.py --keyword "Warranty claim process"
    python run_kb_tests.py --delay 0         # no sleep between calls
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
import time
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

TEST_CASES_PATH = BASE_DIR / "data" / "Test_Cases_MainDB_new.xlsx"
TEST_SHEET      = "Test Cases"
RESULTS_SHEET   = "Results"

RESULT_COLUMNS = [
    "No", "keyword", "question",
    "matched_keyword", "whatsapp_reply",
    "status", "latency_ms",
]

# ── colours ────────────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

def _c(colour: str, text: str) -> str:
    return f"{colour}{text}{RESET}"

def _bar(rate: float, width: int = 20) -> str:
    filled = int(rate / 100 * width)
    return "█" * filled + "░" * (width - filled)


# ── pass / fail ────────────────────────────────────────────────────────────────
def _words(s: str) -> set[str]:
    return {w for w in re.sub(r"\s+", " ", s).lower().split() if len(w) > 2}

def _is_pass(expected: str, got: str) -> bool:
    if not got or got.upper() in ("", "NO_MATCH", "PRODUCT_ENQUIRE", "TICKET_LOGGED"):
        return False
    if re.sub(r"\s+", " ", expected).strip().lower() == re.sub(r"\s+", " ", got).strip().lower():
        return True
    e, g = _words(expected), _words(got)
    return bool(e) and len(e & g) / len(e) >= 0.4


# ── incremental Excel helpers ─────────────────────────────────────────────────
def _init_results_sheet(path: Path) -> None:
    """Create Results sheet with header only if it does not already exist."""
    wb = load_workbook(path)
    if RESULTS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(RESULTS_SHEET)
        ws.append(RESULT_COLUMNS)
        wb.save(path)
    wb.close()


def _clear_results_sheet(path: Path) -> None:
    """Delete and re-create the Results sheet (used by --restart)."""
    wb = load_workbook(path)
    if RESULTS_SHEET in wb.sheetnames:
        del wb[RESULTS_SHEET]
    ws = wb.create_sheet(RESULTS_SHEET)
    ws.append(RESULT_COLUMNS)
    wb.save(path)
    wb.close()


def _load_completed_nos(path: Path) -> set[str]:
    """Return the set of 'No' values already present in the Results sheet."""
    wb = load_workbook(path, read_only=True)
    if RESULTS_SHEET not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[RESULTS_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) <= 1:          # header only
        return set()
    header = [str(c).strip().lower() for c in rows[0]]
    try:
        no_idx = header.index("no")
    except ValueError:
        return set()
    return {str(r[no_idx]).strip() for r in rows[1:] if r[no_idx] is not None}


def _append_result_row(path: Path, row: dict) -> None:
    """Append one result row to Excel immediately."""
    wb = load_workbook(path)
    ws = wb[RESULTS_SHEET]
    ws.append([str(row.get(col, "")) for col in RESULT_COLUMNS])
    wb.save(path)
    wb.close()


def _save_backup(path: Path) -> None:
    stem, suffix, parent = path.stem, path.suffix, path.parent
    n = 1
    while (parent / f"{stem}_backup_{n}{suffix}").exists():
        n += 1
    backup_path = parent / f"{stem}_backup_{n}{suffix}"
    try:
        shutil.copy2(path, backup_path)
        print(f"\nBackup  → {backup_path.name}")
    except Exception as exc:
        print(f"\n[BACKUP ERROR] {exc}")


# ── main ───────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description="Simulate WhatsApp messages and capture replies")
    parser.add_argument("--limit",   type=int,   default=0,     help="Max rows to run this session (0 = all remaining)")
    parser.add_argument("--keyword", type=str,   default="",    help="Filter to one keyword")
    parser.add_argument("--delay",   type=float, default=0.3,   help="Seconds between calls")
    parser.add_argument("--restart", action="store_true",       help="Clear previous results and run all from scratch")
    args = parser.parse_args()

    # ── load test cases ────────────────────────────────────────────────────────
    if not TEST_CASES_PATH.exists():
        print(f"[ERROR] {TEST_CASES_PATH} not found.")
        sys.exit(1)

    df = pd.read_excel(TEST_CASES_PATH, sheet_name=TEST_SHEET)
    df.columns = [c.strip().lower() for c in df.columns]

    if "question" not in df.columns or "keyword" not in df.columns:
        print("[ERROR] Test Cases sheet must have 'question' and 'keyword' columns.")
        sys.exit(1)

    if args.keyword:
        mask = df["keyword"].str.contains(args.keyword, case=False, na=False)
        df   = df[mask].copy()
        print(f"Filtered  → {len(df)} rows matching keyword: {args.keyword!r}")

    # ── prepare Results sheet ──────────────────────────────────────────────────
    if args.restart:
        print(f"\nRestarting — clearing previous results…")
        _clear_results_sheet(TEST_CASES_PATH)
        completed_nos: set[str] = set()
    else:
        _init_results_sheet(TEST_CASES_PATH)
        completed_nos = _load_completed_nos(TEST_CASES_PATH)

    # ── skip already-completed rows ────────────────────────────────────────────
    if completed_nos:
        before = len(df)
        df = df[~df.apply(
            lambda r: str(r.get("no", "")).strip() in completed_nos, axis=1
        )].copy()
        skipped = before - len(df)
        print(f"Resuming  → {skipped} already done, {len(df)} remaining")

    if args.limit and args.limit < len(df):
        df = df.iloc[: args.limit].copy()
        print(f"Limited   → running next {args.limit} rows this session")

    total = len(df)
    if total == 0:
        print("\nAll test cases already completed. Use --restart to re-run.")
        sys.exit(0)

    # ── import search() ────────────────────────────────────────────────────────
    try:
        from chat_services import search
    except Exception as exc:
        print(f"[ERROR] Cannot import search(): {exc}")
        sys.exit(1)

    print(f"\n{BOLD}Simulating {total} WhatsApp messages through the chatbot…{RESET}\n")
    print(f"{'No':>4}  {'Status':<8}  {'ms':>6}  {'Question':<50}  Matched keyword")
    print("─" * 125)

    results_cache: list[dict] = []
    passed = failed = errors = 0

    for loop_idx, (_, row) in enumerate(df.iterrows(), start=1):
        question = str(row.get("question", "")).strip()
        expected = str(row.get("keyword",  "")).strip()
        no       = row.get("no", loop_idx)

        if not question:
            continue

        start_t = time.perf_counter()
        matched_keyword = ""
        whatsapp_reply  = ""
        status          = "ERROR"

        try:
            result = search(
                user_question=question,
                engine_mode=PROVIDER_PRIMARY,
                conversation_summary="",
                conversation_id=None,
            )

            whatsapp_reply  = str(result.get("answer", "")).strip()
            matched_keyword = str(result.get("anchor_token") or "").strip()

            status = "PASS" if _is_pass(expected, matched_keyword) else "FAIL"
            if status == "PASS":
                passed += 1
            else:
                failed += 1

        except Exception as exc:
            matched_keyword = f"EXCEPTION: {exc}"
            status = "ERROR"
            errors += 1

        elapsed_ms = int((time.perf_counter() - start_t) * 1000)

        colour = GREEN if status == "PASS" else (RED if status == "FAIL" else YELLOW)
        print(
            f"{no:>4}  {_c(colour, f'{status:<8}')}  {elapsed_ms:>6} ms  "
            f"{question[:50]:<50}  {matched_keyword[:55]}"
        )
        if status == "FAIL":
            print(
                f"{'':>4}  {'':8}  {'':>6}     "
                f"{'':50}  {_c(YELLOW, 'expected: ')}{expected[:55]}"
            )

        record = {
            "No":              no,
            "keyword":         expected,
            "question":        question,
            "matched_keyword": matched_keyword,
            "whatsapp_reply":  whatsapp_reply,
            "status":          status,
            "latency_ms":      elapsed_ms,
        }
        try:
            _append_result_row(TEST_CASES_PATH, record)
        except Exception as save_exc:
            print(f"  [SAVE ERROR] row {no}: {save_exc}")

        results_cache.append(record)

        if args.delay > 0:
            time.sleep(args.delay)

    # ── summary (this session only) ────────────────────────────────────────────
    saved = len(results_cache)
    print(f"\n{'═' * 70}")
    print(f"{BOLD}SUMMARY  —  {saved} rows run this session{RESET}")
    if saved:
        print(f"  {_c(GREEN,  'PASS')}  : {passed:>4}  ({passed/saved*100:5.1f} %)")
        print(f"  {_c(RED,    'FAIL')}  : {failed:>4}  ({failed/saved*100:5.1f} %)")
    if errors:
        print(f"  {_c(YELLOW, 'ERROR')} : {errors:>4}")

    # ── full sheet summary (all runs combined) ─────────────────────────────────
    try:
        all_results = pd.read_excel(TEST_CASES_PATH, sheet_name=RESULTS_SHEET)
        all_results.columns = [c.strip().lower() for c in all_results.columns]
        total_done = len(all_results)
        total_pass = (all_results["status"] == "PASS").sum()
        print(f"\n  Total in Results sheet: {total_done} rows  |  Overall PASS rate: {total_pass/total_done*100:.1f}%")
    except Exception:
        pass

    print(f"{'═' * 70}")
    print(f"\nResults → {TEST_CASES_PATH}  (sheet: '{RESULTS_SHEET}')")

    # ── per-keyword pass rate (this session) ──────────────────────────────────
    if results_cache:
        results_df = pd.DataFrame(results_cache)
        print(f"\n{BOLD}Per-keyword pass rate (this session):{RESET}")
        breakdown = (
            results_df.groupby("keyword")["status"]
            .value_counts().unstack(fill_value=0)
            .assign(total=lambda d: d.sum(axis=1))
        )
        for col in ("PASS", "FAIL", "ERROR"):
            if col not in breakdown.columns:
                breakdown[col] = 0
        breakdown["pass_rate"] = (breakdown["PASS"] / breakdown["total"] * 100).round(1)
        for kw, r in breakdown.sort_values("pass_rate").iterrows():
            rate   = r["pass_rate"]
            p, f_  = int(r["PASS"]), int(r["FAIL"])
            colour = GREEN if rate >= 80 else (YELLOW if rate >= 50 else RED)
            kw_short = kw.replace("\n", " / ")[:65]
            print(
                f"  {_c(colour, f'{rate:5.1f}%')}  {_bar(rate)}  "
                f"[{p}✓ {f_}✗]  {kw_short}"
            )

    _save_backup(TEST_CASES_PATH)


if __name__ == "__main__":
    main()
